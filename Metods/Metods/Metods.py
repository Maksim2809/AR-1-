import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import random

rd = 'F:\msdownld.tmp\AR.xlsx'
data_xlsx = openpyxl.load_workbook(rd)
book1 = data_xlsx.active
cell_obj = book1.cell(row = 1, column = 1)
m_row = book1.max_row

print('-----')

data = [ 0 for i in range(0,m_row)]

for i in range(1, m_row + 1):
    cell_obj = book1.cell(row = i, column = 1)
    data[i-1] = cell_obj.value

sum_y  = sum(data)
disp = sum([(i-sum_y/m_row)**2  for i in data])/(m_row-1)


print('дисперсия, мат ожидание', disp, sum_y/m_row)

sum_x  = sum([ i for i in range(1,m_row+1)])
sum_xy = sum([x*y for x,y in zip(data,[i for i in range(1,m_row+1)])])
sum_xx = sum([i*i for i in range(1,m_row+1)])

a = (m_row*sum_xy - sum_x*sum_y)/(m_row*sum_xx - (sum_x)**2)
b = (sum_y - a*sum_x)/m_row

print(a,b)
print('------------------------------------')
print('------------------------------------')
ar = data
sum_H  = sum_y
sum_h  = sum_y - data[m_row-1]
sum_Hh = sum([data[i]*data[i-1] for i in range(1,m_row)])
sum_hh = sum([data[i]*data[i] for i in range(0,m_row-1)])

a_1 = (sum_H**2-(m_row)*sum_Hh)/(sum_h*sum_H - m_row*sum_hh)
a_0 = (sum_H - a_1*sum_h)/(m_row)
print(a_0,a_1)
print('------------------------------------')
print('------------------------------------')
AA = sum([data[i]*data[i-1] for i in range(1,m_row)])/sum_hh
print('AA= ', AA)

print('с a0 - ', a_0 + a_1*data[m_row-1])
print('без a0 - ', AA*data[m_row-1])


print('МНК - ',b, a )
print('метод моментов - ', sum_y/m_row, disp)
print('метод МПП с a0', a_0, a_1)
print('метод с a0 -', AA)


e = [random.random() for i in range(0,m_row)]

x1 = [i for i in range(1,m_row+1)]
y1 = data
hk = data[0]
y2 = [hk]
for i in range(1,m_row):
    hk = a_1*hk + e[i]
    y2 = y2+[hk]

print(e)

fig, ax = plt.subplots()
y_max = max(data)
y_min = min(data)
plt.xlim(0,m_row+1)
plt.ylim(y_min-disp-3, y_max+disp+3)
plt.title('AR')
ax.minorticks_on()
ax.grid(which='major',
        color = 'k', 
        linestyle = ':')
ax.grid(which='minor', 
        color = 'k', 
        linestyle = ':')

ax.plot(x1,y1)
ax.plot(x1,y2)

plt.show()
